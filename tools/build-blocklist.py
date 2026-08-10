#!/usr/bin/env python3
"""
build-blocklist.py — regenerate blocklist.csv from the FracTEL outbound rate
deck.

    python3 tools/build-blocklist.py --deck ~/decks/ShortDuration-2026-08.csv

Emits every contiguous-48 prefix rated at or above the threshold (default
$0.10/min) as `prefix,rate_per_min,destination`. Run it after every deck
update; the deck changes monthly and a stale blocklist is a cost control that
has quietly stopped controlling costs.

    --deck PATH        the rate deck. .csv/.tsv natively; .xlsx needs openpyxl.
    --out PATH         output (default blocklist.csv). '-' writes to stdout.
    --threshold RATE   minimum $/min to block (default 0.10).
    --scope SCOPE      conus (default) | nanp | all. See below.
    --code-col NAME    override column autodetection.
    --rate-col NAME
    --dest-col NAME
    --list-columns     print the deck's detected header and exit.
    --allow-empty      permit writing a blocklist with zero rows.
    --quiet            suppress the summary on stderr.

SCOPE
    conus   country code 1, minus every non-contiguous-48 NANP area code:
            Alaska, Hawaii, the territories, Canada, the Caribbean, toll-free,
            premium and the non-geographic service codes. This is the default
            and it is what the rate deck's US pricing actually describes.
            The Caribbean and 900 are already refused a layer earlier by
            BLOCKED_NPAS in config.env, so leaving them out here is not a gap.
    nanp    country code 1, everything, including Canada and the Caribbean.
    all     every prefix in the deck regardless of country code. The SBC is
            NANP-only (NANP_ONLY=true), so non-1 prefixes are unreachable
            anyway and this exists for auditing the deck, not for production.

WHAT THIS DOES NOT SOLVE
    The deck is LRN-rated: the charge follows where a number has been ported,
    not the number dialled. The SBC performs no LNP dip, so it matches the
    dialled number. Numbers natively in an expensive prefix are caught;
    numbers ported into one are not. See README.md.
"""

import argparse
import csv
import io
import os
import re
import sys
from collections import Counter

# ---------------------------------------------------------------------------
# Column autodetection. Decks get re-exported with renamed headers; matching on
# a list of aliases beats hard-coding one vendor's spelling and breaking every
# time somebody changes "Rate" to "Rate/Min".
# ---------------------------------------------------------------------------
CODE_ALIASES = ("code", "prefix", "digits", "dialcode", "dialprefix",
                "destinationcode", "codes", "npanxx", "breakout")
RATE_ALIASES = ("rateperminute", "rate", "ratepermin", "ratemin", "price",
                "priceperminute", "cost", "costperminute", "permin",
                "perminute", "usd", "amount", "interstate")
DEST_ALIASES = ("destination", "description", "name", "label", "region",
                "destinationname", "areaname", "locality", "state", "notes")

# ---------------------------------------------------------------------------
# NANP area codes that are NOT in the contiguous 48.
#
# Anything here is dropped under --scope conus. Kept as one explicit table
# rather than a clever rule, because there is no clever rule: NANP NPA
# assignment is a list, and a wrong guess here either blocks traffic the
# customer is paying to send or fails to block traffic that costs $2/min.
# ---------------------------------------------------------------------------
NON_CONUS_NPAS = set()

# Alaska, Hawaii.
NON_CONUS_NPAS |= {"907", "808"}

# US territories. Ordinary US destinations for dialling purposes, but not
# contiguous-48 and separately rated in most decks.
NON_CONUS_NPAS |= {"787", "939", "340", "671", "670", "684"}

# Canada.
NON_CONUS_NPAS |= {
    "204", "226", "236", "249", "250", "263", "289", "306", "343", "354",
    "365", "367", "368", "382", "387", "403", "416", "418", "428", "431",
    "437", "438", "450", "468", "474", "506", "514", "519", "548", "579",
    "581", "584", "587", "600", "604", "613", "639", "647", "672", "683",
    "705", "709", "742", "753", "778", "780", "782", "807", "819", "825",
    "867", "873", "879", "902", "905",
}

# Caribbean and Atlantic NANP members. The classic IRSF and wangiri
# destinations — inside the NANP, matching _1NXXNXXXXXX perfectly, costing
# dollars a minute.
NON_CONUS_NPAS |= {
    "242", "246", "264", "268", "284", "345", "441", "473", "649", "658",
    "664", "721", "758", "767", "784", "809", "829", "849", "868", "869",
    "876",
}

# Toll-free. Inbound-paid, never a high-cost outbound destination.
NON_CONUS_NPAS |= {"800", "833", "844", "855", "866", "877", "880", "881",
                   "882", "888"}

# Premium and non-geographic service codes.
NON_CONUS_NPAS |= {"900", "500", "521", "522", "523", "524", "525", "526",
                   "527", "528", "529", "532", "533", "544", "566", "577",
                   "588", "710"}


def norm_header(h):
    return re.sub(r"[^a-z0-9]", "", (h or "").strip().lower())


def pick_column(headers, aliases, override=None):
    """Return the index of the first header matching an alias, or None."""
    norm = [norm_header(h) for h in headers]
    if override:
        want = norm_header(override)
        for i, h in enumerate(norm):
            if h == want:
                return i
        raise SystemExit(
            f"build-blocklist.py: ERROR: no column named {override!r} in the deck.\n"
            f"       Columns found: {', '.join(repr(h) for h in headers)}"
        )
    # Exact alias match first, so "rate" beats "interstaterate" when both exist.
    for alias in aliases:
        for i, h in enumerate(norm):
            if h == alias:
                return i
    for alias in aliases:
        for i, h in enumerate(norm):
            if alias in h:
                return i
    return None


def read_rows(path):
    """Yield the deck as a list of rows (list of str). Handles csv/tsv/xlsx."""
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit(
                "build-blocklist.py: ERROR: reading .xlsx needs openpyxl.\n"
                "       Either:  pip3 install openpyxl\n"
                "       Or export the deck as CSV from the portal and pass that."
            )
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
        wb.close()
        return rows

    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        sample = fh.read(64 * 1024)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [row for row in csv.reader(fh, dialect)]


RATE_RE = re.compile(r"-?\d+(?:\.\d+)?")


def parse_rate(raw):
    """'$0.1000 USD' -> 0.1. Returns None if there is no number in there."""
    if raw is None:
        return None
    m = RATE_RE.search(str(raw).replace(",", ""))
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def normalise_prefix(raw):
    """
    '+1-252-761' / '252761' / '1252761' -> ('1252761', True)
    '25212555' (Somalia)                -> ('25212555', False)

    Returns (prefix, is_nanp) or None if there are no digits at all.

    Do NOT blindly prepend a 1 to anything lacking one. Somalia is country
    code 252 and '25212555' would become '125212555' — a syntactically
    plausible US prefix under NPA 252 that is nothing of the sort. A deck
    with international rows in it would quietly seed the blocklist with
    fabricated domestic prefixes.

    A bare NANP prefix is recognised by shape instead: NPA-only (3), NPA+NXX
    partial (6), NPA+NXX (7) or NPA+NXX+line (10), with NPA and NXX both
    starting 2-9, which is the rule NANP itself enforces.
    """
    digits = re.sub(r"\D", "", str(raw or ""))
    if not digits:
        return None

    if digits[0] == "1":
        return (digits, True) if 4 <= len(digits) <= 11 else None

    looks_nanp = (
        len(digits) in (3, 6, 7, 10)
        and digits[0] in "23456789"
        and (len(digits) < 6 or digits[3] in "23456789")
    )
    return ("1" + digits, True) if looks_nanp else (digits, False)


def nanp_reachable(prefix):
    """
    Can this prefix ever match a number the SBC will actually dial?

    Destinations are normalised to 11 digits, 1 + NPA + NXX + line, with NPA
    and NXX both starting 2-9. A prefix that violates that can never match, so
    emitting it is dead weight in the generated dialplan and, worse, reads
    like coverage that does not exist.
    """
    if not prefix.startswith("1") or not (4 <= len(prefix) <= 11):
        return False
    if prefix[1] not in "23456789":
        return False
    if len(prefix) >= 5 and prefix[4] not in "23456789":
        return False
    return True


def sanitise_dest(raw):
    """
    Strip anything that would break the generated dialplan or the CSV.

    The label is interpolated into a Set() in extensions.conf, where a comma
    is an argument separator, and into a WARN line that gets grepped. Neither
    survives free-form carrier text intact.
    """
    s = re.sub(r'[",;|]', " ", str(raw or ""))
    s = re.sub(r"\s+", " ", s).strip()
    return s[:40]


def main():
    ap = argparse.ArgumentParser(add_help=True, description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--deck", required=True)
    ap.add_argument("--out", default="blocklist.csv")
    ap.add_argument("--threshold", type=float, default=0.10)
    ap.add_argument("--scope", choices=("conus", "nanp", "all"), default="conus")
    ap.add_argument("--code-col")
    ap.add_argument("--rate-col")
    ap.add_argument("--dest-col")
    ap.add_argument("--list-columns", action="store_true")
    ap.add_argument("--allow-empty", action="store_true")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.deck):
        raise SystemExit(f"build-blocklist.py: ERROR: no such deck: {args.deck}")

    rows = read_rows(args.deck)
    if not rows:
        raise SystemExit(f"build-blocklist.py: ERROR: {args.deck} is empty")

    # The header is not always row 1 — decks often carry a title block above
    # it. Take the first row in which a code column and a rate column are both
    # detectable.
    header_idx, headers = None, None
    for i, row in enumerate(rows[:25]):
        if pick_column(row, CODE_ALIASES) is not None and \
           pick_column(row, RATE_ALIASES) is not None:
            header_idx, headers = i, row
            break
    if headers is None:
        raise SystemExit(
            "build-blocklist.py: ERROR: could not find a header row with both a\n"
            "       code column and a rate column in the first 25 rows of\n"
            f"       {args.deck}\n"
            f"       First row seen: {rows[0]}\n"
            "       Name them explicitly with --code-col and --rate-col, or run\n"
            "       --list-columns to see what is there."
        )

    if args.list_columns:
        print(f"header row {header_idx + 1} of {args.deck}:")
        for i, h in enumerate(headers):
            print(f"  [{i}] {h!r}")
        return 0

    ci = pick_column(headers, CODE_ALIASES, args.code_col)
    ri = pick_column(headers, RATE_ALIASES, args.rate_col)
    di = pick_column(headers, DEST_ALIASES, args.dest_col)

    best = {}          # prefix -> (rate, destination)
    stats = Counter()
    dropped_npas = Counter()

    for row in rows[header_idx + 1:]:
        if not row or all(not str(c).strip() for c in row):
            continue
        stats["read"] += 1

        parsed = normalise_prefix(row[ci] if ci < len(row) else "")
        rate = parse_rate(row[ri] if ri < len(row) else "")

        if parsed is None:
            stats["bad_prefix"] += 1
            continue
        prefix, is_nanp = parsed
        if rate is None:
            stats["bad_rate"] += 1
            continue

        if args.scope != "all" and not is_nanp:
            stats["out_of_scope"] += 1
            continue
        if args.scope == "conus":
            npa = prefix[1:4]
            if len(npa) == 3 and npa in NON_CONUS_NPAS:
                stats["out_of_scope"] += 1
                dropped_npas[npa] += 1
                continue

        # A prefix the dialplan can never match is not coverage. Drop it and
        # say so rather than padding the file with rows that do nothing.
        if not nanp_reachable(prefix):
            stats["unreachable"] += 1
            continue

        stats["in_scope"] += 1
        if rate < args.threshold:
            stats["below_threshold"] += 1
            continue

        dest = sanitise_dest(row[di] if (di is not None and di < len(row)) else "")
        # The same prefix can appear more than once in a deck (time-of-day or
        # inter/intrastate splits). Keep the WORST rate: blocking on the
        # cheapest variant would let the expensive one through.
        prev = best.get(prefix)
        if prev is None or rate > prev[0]:
            best[prefix] = (rate, dest)

    if not best and not args.allow_empty:
        raise SystemExit(
            f"build-blocklist.py: ERROR: nothing at or above ${args.threshold:.4f}/min\n"
            f"       in scope '{args.scope}' - refusing to write an empty blocklist\n"
            "       over your existing one. That would silently disable the primary\n"
            "       cost control.\n"
            f"       Read {stats['read']} row(s), {stats['in_scope']} in scope.\n"
            "       Check --threshold, --scope, and that the rate column was\n"
            "       detected correctly (--list-columns)."
        )

    out = io.StringIO()
    w = csv.writer(out, lineterminator="\n")
    w.writerow(["prefix", "rate_per_min", "destination"])
    # Sorted so a regenerated file diffs cleanly against the previous month.
    for prefix in sorted(best):
        rate, dest = best[prefix]
        w.writerow([prefix, f"{rate:.6f}", dest])
    text = out.getvalue()

    if args.out == "-":
        sys.stdout.write(text)
    else:
        tmp = args.out + ".tmp"
        with open(tmp, "w", encoding="utf-8", newline="") as fh:
            fh.write(text)
        os.replace(tmp, args.out)

    if not args.quiet:
        rates = [r for r, _ in best.values()]
        e = sys.stderr.write
        e(f"\nbuild-blocklist.py - {args.deck}\n")
        e(f"  header row            {header_idx + 1}\n")
        e(f"  code column           {headers[ci]!r}\n")
        e(f"  rate column           {headers[ri]!r}\n")
        if di is not None:
            e(f"  destination column    {headers[di]!r}\n")
        else:
            e("  destination column    (none found — labels will be blank)\n")
        e(f"  scope                 {args.scope}\n")
        e(f"  threshold             ${args.threshold:.4f}/min\n")
        e("  " + "-" * 52 + "\n")
        e(f"  rows read             {stats['read']}\n")
        e(f"  unparseable prefix    {stats['bad_prefix']}\n")
        e(f"  unparseable rate      {stats['bad_rate']}\n")
        e(f"  out of scope          {stats['out_of_scope']}\n")
        e(f"  in scope              {stats['in_scope']}\n")
        e(f"  below threshold       {stats['below_threshold']}\n")
        e(f"  unreachable pattern   {stats['unreachable']}\n")
        e(f"  BLOCKED               {len(best)}\n")
        if rates:
            e(f"    at or above $0.50   {sum(1 for r in rates if r >= 0.50)}\n")
            e(f"    at or above $1.00   {sum(1 for r in rates if r >= 1.00)}\n")
            e(f"    worst rate          ${max(rates):.6f}/min\n")
        if dropped_npas:
            top = ", ".join(f"{n}({c})" for n, c in dropped_npas.most_common(8))
            e(f"  non-CONUS NPAs dropped: {top}\n")
            e("    (the Caribbean and 900 are refused a layer earlier by\n"
              "     BLOCKED_NPAS in config.env - not a gap)\n")
        if args.out != "-":
            e(f"\n  wrote {args.out}\n")
            e("  Now re-render and reload:  sudo ./install.sh --config-only\n\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
